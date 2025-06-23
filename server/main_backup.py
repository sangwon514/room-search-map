from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel
from typing import List, Optional
import uvicorn
import os
import asyncio
import httpx
from datetime import datetime, timedelta
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from urllib.parse import quote

# 환경 변수 로드
load_dotenv()

# FastAPI 앱 생성
app = FastAPI(
    title="Room Crawler API",
    description="숙소 검색 및 예약률 관리 API",
    version="1.0.0"
)

# CORS 설정
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:5174"],  # React 개발 서버
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 데이터 모델 정의
class Room(BaseModel):
    rid: int
    room_name: str
    state: str
    province: str
    town: str
    pic_main: str
    addr_lot: str
    addr_street: str
    using_fee: int
    pyeong_size: float
    room_cnt: int
    bathroom_cnt: int
    cookroom_cnt: int
    sittingroom_cnt: int
    reco_type_1: bool
    reco_type_2: bool
    longterm_discount_per: int
    early_discount_per: int
    is_new: bool
    is_super_host: bool
    lat: float
    lng: float

class SearchParams(BaseModel):
    keyword: Optional[str] = ""
    theme_type: Optional[str] = ""
    room_cnt: Optional[List[str]] = []
    property_type: Optional[List[str]] = []
    animal: Optional[bool] = False
    subway: Optional[bool] = False
    longterm_discount: Optional[bool] = False
    early_discount: Optional[bool] = False
    parking_place: Optional[bool] = False
    min_using_fee: Optional[int] = 0
    max_using_fee: Optional[int] = 1000000
    sort: Optional[str] = "latest"
    now_page: Optional[int] = 1
    itemcount: Optional[int] = 20
    by_location: Optional[bool] = False
    north_east_lat: Optional[float] = 0
    north_east_lng: Optional[float] = 0
    south_west_lat: Optional[float] = 0
    south_west_lng: Optional[float] = 0
    map_level: Optional[int] = 3

class RoomListResponse(BaseModel):
    list: List[Room]
    total_count: int
    current_page: int

# 예약률 관련 모델 추가
class ReservationRequest(BaseModel):
    rid_list: List[int]
    start_year: int
    start_month: int
    end_year: int
    end_month: int

class ScheduleItem(BaseModel):
    date: str
    status: str

class ScheduleResponse(BaseModel):
    error_code: Optional[int] = None
    schedule_list: Optional[List[ScheduleItem]] = []

class ReservationData(BaseModel):
    rid: int
    year: int
    month: int
    schedule_list: List[ScheduleItem] = []
    error_code: Optional[int] = 0
    raw_response: Optional[dict] = None

class ReservationBatchResponse(BaseModel):
    success: bool
    total_requests: int
    completed_requests: int
    failed_requests: int
    data: List[ReservationData]
    errors: List[str]

# 세션 확인 헬퍼 함수
def get_session_from_cookies(request: Request) -> Optional[str]:
    """쿠키에서 세션 값을 추출"""
    session = request.cookies.get("session")
    return session

# API 엔드포인트
@app.get("/")
async def root():
    """API 상태 확인"""
    return {"message": "Room Crawler API is running", "version": "1.0.0"}

@app.get("/health")
async def health_check():
    """헬스 체크"""
    return {"status": "healthy", "timestamp": "2025-06-23T15:49:00+09:00"}

@app.post("/api/rooms", response_model=RoomListResponse)
async def search_rooms(search_params: SearchParams, request: Request):
    """숙소 검색 API"""
    # 세션 확인
    session = get_session_from_cookies(request)
    if not session:
        raise HTTPException(status_code=401, detail="세션이 설정되지 않았습니다.")
    
    # TODO: 실제 데이터베이스 또는 외부 API 연동
    # 현재는 더미 데이터 반환
    dummy_rooms = [
        Room(
            rid=1,
            room_name="신촌 원룸",
            state="서울",
            province="서대문구",
            town="신촌동",
            pic_main="https://example.com/image1.jpg",
            addr_lot="서울 서대문구 신촌동 123-45",
            addr_street="서울 서대문구 연세로 50",
            using_fee=500000,
            pyeong_size=10.5,
            room_cnt=1,
            bathroom_cnt=1,
            cookroom_cnt=1,
            sittingroom_cnt=1,
            reco_type_1=True,
            reco_type_2=False,
            longterm_discount_per=10,
            early_discount_per=5,
            is_new=True,
            is_super_host=False,
            lat=37.5665,
            lng=126.9365
        ),
        Room(
            rid=2,
            room_name="홍대 투룸",
            state="서울",
            province="마포구",
            town="홍익동",
            pic_main="https://example.com/image2.jpg",
            addr_lot="서울 마포구 홍익동 456-78",
            addr_street="서울 마포구 홍익로 100",
            using_fee=700000,
            pyeong_size=15.2,
            room_cnt=2,
            bathroom_cnt=1,
            cookroom_cnt=1,
            sittingroom_cnt=1,
            reco_type_1=False,
            reco_type_2=True,
            longterm_discount_per=15,
            early_discount_per=8,
            is_new=False,
            is_super_host=True,
            lat=37.5563,
            lng=126.9236
        )
    ]
    
    return RoomListResponse(
        list=dummy_rooms,
        total_count=len(dummy_rooms),
        current_page=search_params.now_page or 1
    )

@app.get("/api/session")
async def get_session_info(request: Request):
    """현재 세션 정보 조회"""
    session = get_session_from_cookies(request)
    if not session:
        raise HTTPException(status_code=401, detail="세션이 설정되지 않았습니다.")
    
    return {"session": session, "status": "active"}

@app.post("/api/reservations", response_model=ReservationBatchResponse)
async def get_reservations(reservation_request: ReservationRequest, request: Request):
    """예약률 데이터 가져오기 API"""
    print(f"🚀 [DEBUG] 예약률 API 요청 시작")
    print(f"  - RID 목록: {reservation_request.rid_list}")
    print(f"  - 기간: {reservation_request.start_year}-{reservation_request.start_month:02d} ~ {reservation_request.end_year}-{reservation_request.end_month:02d}")
    
    # 세션 확인
    session = get_session_from_cookies(request)
    print(f"  - 세션 추출: {'성공' if session else '실패'}")
    if session:
        print(f"  - 세션 값: {session[:20]}..." if len(session) > 20 else f"  - 세션 값: {session}")
    
    if not session:
        print(f"❌ [DEBUG] 세션이 없어서 401 반환")
        raise HTTPException(status_code=401, detail="세션이 설정되지 않았습니다.")
    
    # 요청할 년월 목록 생성
    year_month_list = []
    start_date = datetime(reservation_request.start_year, reservation_request.start_month, 1)
    end_date = datetime(reservation_request.end_year, reservation_request.end_month, 1)
    
    current_date = start_date
    while current_date <= end_date:
        year_month_list.append((current_date.year, current_date.month))
        # 다음 달로 이동
        if current_date.month == 12:
            current_date = current_date.replace(year=current_date.year + 1, month=1)
        else:
            current_date = current_date.replace(month=current_date.month + 1)
    
    print(f"  - 생성된 년월 목록: {year_month_list}")
    
    # 모든 요청 조합 생성 (rid * year_month)
    all_requests = []
    for rid in reservation_request.rid_list:
        for year, month in year_month_list:
            all_requests.append({
                "rid": rid,
                "year": year,
                "month": month
            })
    
    print(f"  - 총 요청 수: {len(all_requests)}")
    
    # 결과 저장
    reservation_data = []
    errors = []
    completed_requests = 0
    
    # 10개씩 배치로 처리
    batch_size = 10
    external_api_url = "https://33m2.co.kr/app/room/schedule"
    
    print(f"🔄 [DEBUG] 배치 처리 시작 (배치 크기: {batch_size})")
    
    async with httpx.AsyncClient(timeout=30.0) as client:
        for i in range(0, len(all_requests), batch_size):
            batch = all_requests[i:i + batch_size]
            batch_num = (i // batch_size) + 1
            total_batches = (len(all_requests) + batch_size - 1) // batch_size
            
            print(f"📦 [DEBUG] 배치 {batch_num}/{total_batches} 처리 중 ({len(batch)}개 요청)")
            
            # 배치 내 요청들을 병렬로 처리
            tasks = [
                fetch_schedule_data(client, external_api_url, session, req["rid"], req["year"], req["month"])
                for req in batch
            ]
            results = await asyncio.gather(*tasks, return_exceptions=True)
            
            # 결과 처리
            for j, result in enumerate(results):
                if isinstance(result, Exception):
                    error_msg = f"RID {batch[j]['rid']} ({batch[j]['year']}-{batch[j]['month']:02d}) 실패: {str(result)}"
                    errors.append(error_msg)
                    print(f"❌ [DEBUG] {error_msg}")
                    
                    # 403 오류인 경우 즉시 중단
                    if "403" in str(result):
                        print(f"🚨 [DEBUG] 403 오류로 인한 즉시 중단!")
                        return ReservationBatchResponse(
                            success=False,
                            total_requests=len(all_requests),
                            completed_requests=completed_requests,
                            failed_requests=len(errors),
                            data=reservation_data,
                            errors=errors
                        )
                else:
                    # result는 이미 ReservationData 객체임
                    reservation_data.append(result)
                    completed_requests += 1
                    print(f"✅ [DEBUG] RID {batch[j]['rid']} ({batch[j]['year']}-{batch[j]['month']:02d}) 성공")
            
            # 다음 배치 전 1초 대기 (마지막 배치가 아닌 경우)
            if i + batch_size < len(all_requests):
                print(f"⏳ [DEBUG] 다음 배치 전 1초 대기...")
                await asyncio.sleep(1)
    
    print(f"🏁 [DEBUG] 모든 배치 처리 완료")
    print(f"  - 성공: {completed_requests}/{len(all_requests)}")
    print(f"  - 실패: {len(errors)}")
    
    return ReservationBatchResponse(
        success=len(errors) == 0,
        total_requests=len(all_requests),
        completed_requests=completed_requests,
        failed_requests=len(errors),
        data=reservation_data,
        errors=errors
    )

async def fetch_schedule_data(client: httpx.AsyncClient, url: str, session: str, rid: int, year: int, month: int):
    """외부 API에서 스케줄 데이터 가져오기"""
    payload = {
        "rid": str(rid),  # 문자열로 변환
        "year": str(year),  # 문자열로 변환
        "month": str(month).zfill(2)  # 2자리 문자열로 변환 (07)
    }
    
    headers = {
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/137.0.0.0 Safari/537.36",
        "Referer": f"https://33m2.co.kr/room/detail/{rid}",
        "Origin": "https://33m2.co.kr",
        "x-requested-with": "XMLHttpRequest",
        "Cookie": f"SESSION={session}"
    }
    
    print(f"🔍 [DEBUG] 요청 정보:")
    print(f"  - URL: {url}")
    print(f"  - RID: {rid}, Year: {year}, Month: {month}")
    print(f"  - Session: {session[:20]}..." if len(session) > 20 else f"  - Session: {session}")
    print(f"  - Headers: {headers}")
    print(f"  - Payload: {payload}")
    
    try:
        response = await client.post(url, data=payload, headers=headers)
        
        print(f"📡 [DEBUG] 응답 정보:")
        print(f"  - Status Code: {response.status_code}")
        print(f"  - Response Headers: {dict(response.headers)}")
        print(f"  - Response Text (first 200 chars): {response.text[:200]}...")
        
        if response.status_code == 403:
            print(f"❌ [DEBUG] 403 Forbidden 오류 발생!")
            print(f"  - 전체 응답 내용: {response.text}")
            raise HTTPException(status_code=403, detail=f"외부 API 403 오류: {response.text}")
        
        response.raise_for_status()
        
        # JSON 응답 파싱
        json_data = response.json()
        print(f"✅ [DEBUG] JSON 파싱 성공: {type(json_data)}")
        print(f"  - error_code: {json_data.get('error_code')}")
        print(f"  - schedule_list 개수: {len(json_data.get('schedule_list', []))}")
        
        # schedule_list 파싱
        schedule_list = []
        if "schedule_list" in json_data and isinstance(json_data["schedule_list"], list):
            for item in json_data["schedule_list"]:
                if isinstance(item, dict) and "date" in item and "status" in item:
                    schedule_list.append(ScheduleItem(
                        date=item["date"],
                        status=item["status"]
                    ))
        
        # 응답 데이터에 요청 정보 추가
        result = {
            "rid": rid,
            "year": year,
            "month": month,
            "schedule_list": schedule_list,
            "error_code": json_data.get("error_code", 0),
            "raw_response": json_data
        }
        
        print(f"📊 [DEBUG] 파싱된 데이터:")
        print(f"  - schedule_list 개수: {len(schedule_list)}")
        print(f"  - error_code: {result['error_code']}")
        
        return ReservationData(**result)
        
    except httpx.HTTPStatusError as e:
        error_msg = f"HTTP 오류 {e.response.status_code}: {e.response.text}"
        print(f"❌ [DEBUG] HTTP 상태 오류: {error_msg}")
        raise Exception(error_msg)
    except httpx.RequestError as e:
        error_msg = f"요청 오류: {str(e)}"
        print(f"❌ [DEBUG] 요청 오류: {error_msg}")
        raise Exception(error_msg)
    except Exception as e:
        error_msg = f"예상치 못한 오류: {str(e)}"
        print(f"❌ [DEBUG] 예상치 못한 오류: {error_msg}")
        raise Exception(error_msg)

# 엑셀 다운로드 기능 추가
@app.post("/api/download_excel")
async def download_excel(reservation_request: ReservationRequest, request: Request):
    """예약률 데이터를 엑셀 파일로 다운로드"""
    print(f"📊 [DEBUG] 엑셀 다운로드 요청 시작")
    
    # 세션 확인
    session = get_session_from_cookies(request)
    if not session:
        raise HTTPException(status_code=401, detail="세션이 필요합니다. 쿠키에 session 값을 설정해주세요.")
    
    # 예약률 데이터 가져오기 (기존 로직 재사용)
    external_api_url = "https://33m2.co.kr/app/room/schedule"
    
    # 요청 목록 생성
    all_requests = []
    for rid in reservation_request.rid_list:
        for year in range(reservation_request.start_year, reservation_request.end_year + 1):
            start_month = reservation_request.start_month if year == reservation_request.start_year else 1
            end_month = reservation_request.end_month if year == reservation_request.end_year else 12
            
            for month in range(start_month, end_month + 1):
                all_requests.append({
                    "rid": rid,
                    "year": year,
                    "month": month
                })
    
    print(f"📋 [DEBUG] 총 {len(all_requests)}개 요청 생성")
    
    # 데이터 수집
    reservation_data = []
    errors = []
    completed_requests = 0
    batch_size = 10
    
    async with httpx.AsyncClient(timeout=30.0) as client:
        for i in range(0, len(all_requests), batch_size):
            batch = all_requests[i:i + batch_size]
            batch_num = (i // batch_size) + 1
            total_batches = (len(all_requests) + batch_size - 1) // batch_size
            
            print(f"📦 [DEBUG] 배치 {batch_num}/{total_batches} 처리 중 ({len(batch)}개 요청)")
            
            # 배치 내 요청들을 병렬로 처리
            tasks = [
                fetch_schedule_data(client, external_api_url, session, req["rid"], req["year"], req["month"])
                for req in batch
            ]
            results = await asyncio.gather(*tasks, return_exceptions=True)
            
            # 결과 처리
            for j, result in enumerate(results):
                if isinstance(result, Exception):
                    error_msg = f"RID {batch[j]['rid']} ({batch[j]['year']}-{batch[j]['month']:02d}) 실패: {str(result)}"
                    errors.append(error_msg)
                    print(f"❌ [DEBUG] {error_msg}")
                    
                    # 403 오류인 경우 즉시 중단
                    if "403" in str(result):
                        print(f"🚨 [DEBUG] 403 오류로 인한 즉시 중단!")
                        raise HTTPException(status_code=403, detail="세션이 유효하지 않거나 권한이 없습니다.")
                else:
                    # result는 이미 ReservationData 객체임
                    reservation_data.append(result)
                    completed_requests += 1
                    print(f"✅ [DEBUG] RID {batch[j]['rid']} ({batch[j]['year']}-{batch[j]['month']:02d}) 성공")
            
            # 다음 배치 전 1초 대기 (마지막 배치가 아닌 경우)
            if i + batch_size < len(all_requests):
                print(f"⏳ [DEBUG] 다음 배치 전 1초 대기...")
                await asyncio.sleep(1)
    
    print(f"🏁 [DEBUG] 데이터 수집 완료. 엑셀 파일 생성 시작...")
    
    # 엑셀 파일 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "예약률 데이터"
    
    # 헤더 행 추가
    headers = ["RID", "숙소명", "년도", "월", "날짜", "예약상태", "요일"]
    ws.append(headers)
    
    # 헤더 스타일 적용
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws.column_dimensions[chr(64 + col_num)].width = 12
    
    # 데이터 행 추가
    row_num = 2
    for data in reservation_data:
        for schedule_item in data.schedule_list:
            # 날짜 파싱
            date_str = schedule_item.date
            try:
                date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                weekday = ["월", "화", "수", "목", "금", "토", "일"][date_obj.weekday()]
            except:
                weekday = ""
            
            # 상태 한글화
            status_map = {
                "disable": "예약불가",
                "enable": "예약가능", 
                "booked": "예약완료",
                "blocked": "차단됨"
            }
            status_kr = status_map.get(schedule_item.status, schedule_item.status)
            
            row_data = [
                data.rid,
                f"숙소_{data.rid}",  # 실제 숙소명은 별도 API에서 가져와야 함
                data.year,
                data.month,
                date_str,
                status_kr,
                weekday
            ]
            ws.append(row_data)
            
            # 데이터 행 스타일
            for col_num, cell in enumerate(ws[row_num], 1):
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if schedule_item.status == "disable":
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                elif schedule_item.status == "booked":
                    cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            
            row_num += 1
    
    # 요약 정보 추가
    ws.append([])
    ws.append(["=== 요약 정보 ==="])
    ws.append([f"총 요청 수: {len(all_requests)}"])
    ws.append([f"성공 요청: {completed_requests}"])
    ws.append([f"실패 요청: {len(errors)}"])
    ws.append([f"데이터 행 수: {row_num - 2}"])
    ws.append([f"생성 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    
    # 엑셀 파일을 메모리에 저장
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # 파일명 생성
    filename = f"예약률_데이터_{reservation_request.start_year}{reservation_request.start_month:02d}_{reservation_request.end_year}{reservation_request.end_month:02d}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    encoded_filename = quote(filename)
    
    print(f"✅ [DEBUG] 엑셀 파일 생성 완료: {filename}")
    
    # 파일 다운로드 응답
    return StreamingResponse(
        BytesIO(buffer.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )

# 예외 처리
@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    return JSONResponse(
        status_code=exc.status_code,
        content={"detail": exc.detail, "status_code": exc.status_code}
    )

if __name__ == "__main__":
    uvicorn.run(
        "main:app",
        host="0.0.0.0",
        port=8000,
        reload=True,
        log_level="info"
    )
