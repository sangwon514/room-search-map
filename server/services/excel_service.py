from datetime import datetime
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from urllib.parse import quote
from collections import defaultdict

from models import ReservationBatchResponse, ReservationRequest

class ExcelService:
    def __init__(self):
        # 스타일 정의
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.center_alignment = Alignment(horizontal="center", vertical="center")
        self.total_font = Font(bold=True, size=12)
        self.total_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    
    def create_excel_file(self, data: ReservationBatchResponse, reservation_request: ReservationRequest) -> BytesIO:
        """월별 예약률 엑셀 파일 생성"""
        wb = Workbook()
        ws = wb.active
        ws.title = "월별 예약률"
        
        # 월 범위 생성
        months = self._get_month_range(reservation_request)
        
        # 헤더 생성
        headers = ["방 ID", "방 이름"] + [f"{month['year']}년 {month['month']}월" for month in months] + ["예약률", "URL"]
        
        # 헤더 스타일 설정
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = self.center_alignment
        
        # 방별 예약 데이터 집계
        room_reservations = self._aggregate_monthly_reservations(data, reservation_request)
        
        # 데이터 행 추가
        row = 2
        total_reserved_days = 0
        total_possible_days = 0
        
        # 현재 날짜 정보
        today = datetime.now()
        current_year = today.year
        current_month = today.month
        current_day = today.day
        
        for room_info in reservation_request.room_list:
            room_key = f"{room_info.rid}_{room_info.rname}"
            monthly_data = room_reservations.get(room_key, {})
            
            # 방 ID와 방 이름 입력
            ws.cell(row=row, column=1, value=room_info.rid)
            ws.cell(row=row, column=2, value=room_info.rname)
            
            room_reserved_days = 0
            room_possible_days = 0
            
            # 각 월별 데이터 입력
            for col, month in enumerate(months, 3):  # 3번째 컬럼부터 시작 (방 ID, 방 이름 다음)
                month_key = f"{month['year']}-{month['month']:02d}"
                reserved_days = monthly_data.get(month_key, 0)
                
                # 해당 월의 총 일수 계산
                if month['year'] == current_year and month['month'] == current_month:
                    # 현재 달: 오늘부터 월말까지
                    days_in_month = calendar.monthrange(month['year'], month['month'])[1]
                    possible_days = days_in_month - current_day + 1
                else:
                    # 과거/미래 달: 전체 일수
                    possible_days = calendar.monthrange(month['year'], month['month'])[1]
                
                ws.cell(row=row, column=col, value=reserved_days)
                room_reserved_days += reserved_days
                room_possible_days += possible_days
                
                total_reserved_days += reserved_days
                total_possible_days += possible_days
            
            # 방별 예약률 계산 및 표시
            room_rate = (room_reserved_days / room_possible_days * 100) if room_possible_days > 0 else 0
            ws.cell(row=row, column=len(headers) - 1, value=f"{room_rate:.1f}%")
            
            # URL 입력
            room_url = f"https://33m2.co.kr/room/detail/{room_info.rid}"
            ws.cell(row=row, column=len(headers), value=room_url)
            
            row += 1
        
        # 빈 행 추가
        row += 1
        
        # 총 예약률 계산
        total_rate = (total_reserved_days / total_possible_days * 100) if total_possible_days > 0 else 0
        
        # 총 예약률 행 추가
        total_cell = ws.cell(row=row, column=1, value="총 예약률")
        total_cell.font = self.total_font
        total_cell.fill = self.total_fill
        
        # 총 예약 일수들 입력
        for col, month in enumerate(months, 3):
            month_key = f"{month['year']}-{month['month']:02d}"
            month_total = sum(monthly_data.get(month_key, 0) for monthly_data in room_reservations.values())
            cell = ws.cell(row=row, column=col, value=month_total)
            cell.font = self.total_font
            cell.fill = self.total_fill
            cell.alignment = self.center_alignment
        
        # 총 예약률 표시
        total_rate_cell = ws.cell(row=row, column=len(headers) - 1, value=f"{total_rate:.1f}%")
        total_rate_cell.font = self.total_font
        total_rate_cell.fill = self.total_fill
        total_rate_cell.alignment = self.center_alignment
        
        # URL 컬럼은 빈 값으로 처리
        url_cell = ws.cell(row=row, column=len(headers), value="")
        url_cell.font = self.total_font
        url_cell.fill = self.total_fill
        
        # 요약 정보 추가
        row += 2
        summary_data = [
            ["생성 시간", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["조회 기간", f"{reservation_request.start_year}년 {reservation_request.start_month}월 ~ {reservation_request.end_year}년 {reservation_request.end_month}월"],
            ["총 방 수", len(room_reservations)],
            ["총 요청 수", data.total_requests],
            ["성공 요청", data.completed_requests],
            ["실패 요청", data.failed_requests],
            ["총 예약 일수", total_reserved_days],
            ["총 가능 일수", total_possible_days],
        ]
        
        for summary_row in summary_data:
            ws.cell(row=row, column=1, value=summary_row[0]).font = Font(bold=True)
            ws.cell(row=row, column=2, value=summary_row[1])
            row += 1
        
        # 컬럼 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 파일을 메모리에 저장
        file_buffer = BytesIO()
        wb.save(file_buffer)
        file_buffer.seek(0)
        
        return file_buffer
    
    def _aggregate_monthly_reservations(self, data: ReservationBatchResponse, reservation_request: ReservationRequest) -> dict:
        """월별 예약 데이터 집계"""
        room_reservations = defaultdict(lambda: defaultdict(int))
        
        # 요청된 모든 RID에 대해 방 이름 초기화
        for room_info in reservation_request.room_list:
            room_key = f"{room_info.rid}_{room_info.rname}"
            room_reservations[room_key] = defaultdict(int)
        
        print(f"📊 [EXCEL] 요청된 방 목록: {[room_info.rname for room_info in reservation_request.room_list]}")
        print(f"📊 [EXCEL] 수집된 데이터 개수: {len(data.data)}")
        
        # 현재 날짜 정보
        today = datetime.now()
        current_year = today.year
        current_month = today.month
        current_day = today.day
        
        # 성공/실패 통계
        success_count = 0
        error_count = 0
        
        for reservation_data in data.data:
            rid = reservation_data.rid
            
            # 요청된 방 목록에서 해당 rid의 방 이름 찾기
            room_info = next((room for room in reservation_request.room_list if room.rid == rid), None)
            if not room_info:
                continue  # 요청되지 않은 방은 스킵
                
            room_key = f"{rid}_{room_info.rname}"
            
            if reservation_data.error_code == 0:  # 성공한 데이터
                success_count += 1
                # 예약된 날짜 카운트 (booking, booked 상태만)
                for schedule_item in reservation_data.schedule_list:
                    if schedule_item.status in ['disable', 'booking']:
                        # 날짜 파싱 (YYYY-MM-DD 형식 가정)
                        try:
                            item_date = datetime.strptime(schedule_item.date, "%Y-%m-%d")
                            
                            # 현재 달인 경우 오늘 이후 날짜만 카운트
                            if (reservation_data.year == current_year and 
                                reservation_data.month == current_month):
                                if item_date.day >= current_day:
                                    month_key = f"{reservation_data.year}-{reservation_data.month:02d}"
                                    room_reservations[room_key][month_key] += 1
                            else:
                                # 과거/미래 달은 모든 날짜 카운트
                                month_key = f"{reservation_data.year}-{reservation_data.month:02d}"
                                room_reservations[room_key][month_key] += 1
                        except ValueError:
                            # 날짜 파싱 실패 시 기존 로직 사용
                            month_key = f"{reservation_data.year}-{reservation_data.month:02d}"
                            room_reservations[room_key][month_key] += 1
                
                print(f"✅ [EXCEL] 방 {rid}({room_info.rname}) ({reservation_data.year}-{reservation_data.month:02d}): {len(reservation_data.schedule_list)}개 일정")
            else:  # 실패한 데이터
                error_count += 1
                print(f"❌ [EXCEL] 방 {rid}({room_info.rname}) ({reservation_data.year}-{reservation_data.month:02d}): 오류 코드 {reservation_data.error_code}")
        
        print(f"📊 [EXCEL] 집계 완료 - 성공: {success_count}, 실패: {error_count}")
        print(f"📊 [EXCEL] 최종 방 개수: {len(room_reservations)}")
        
        return dict(room_reservations)
    
    def _get_month_range(self, reservation_request: ReservationRequest) -> list:
        """요청 기간의 월 목록 생성"""
        months = []
        current_year = reservation_request.start_year
        current_month = reservation_request.start_month
        
        while (current_year < reservation_request.end_year or 
               (current_year == reservation_request.end_year and current_month <= reservation_request.end_month)):
            months.append({"year": current_year, "month": current_month})
            
            current_month += 1
            if current_month > 12:
                current_month = 1
                current_year += 1
        
        return months
    
    def generate_filename(self, reservation_request: ReservationRequest) -> str:
        """파일명 생성"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"월별예약률_{reservation_request.start_year}{reservation_request.start_month:02d}_{reservation_request.end_year}{reservation_request.end_month:02d}_{timestamp}.xlsx"
